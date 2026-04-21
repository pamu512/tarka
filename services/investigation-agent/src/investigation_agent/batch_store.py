"""In-memory tabular batch jobs (CSV / JSON / Excel) for copilot analysis tools."""

from __future__ import annotations

import csv
import io
import json
import os
import re
import threading
import time
import uuid
from collections import Counter
from pathlib import Path
from typing import Any

_MAX_FILE_BYTES = 15 * 1024 * 1024
_MAX_ROWS = 8_000
_MAX_COLS = 128
_DEFAULT_TTL_SECONDS = 2 * 3600
_MAX_STORE_BATCHES = 200


def ttl_seconds() -> int:
    """In-memory batch job retention (env **BATCH_TTL_SECONDS**); clamped 5m–24h."""
    try:
        v = int(os.environ.get("BATCH_TTL_SECONDS", str(_DEFAULT_TTL_SECONDS)))
        return max(300, min(v, 86400))
    except ValueError:
        return _DEFAULT_TTL_SECONDS


_SAFE_BATCH_ID = re.compile(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", re.IGNORECASE)


def validate_batch_id(batch_id: str) -> str:
    bid = str(batch_id).strip()
    if not _SAFE_BATCH_ID.match(bid):
        raise ValueError("Invalid batch_id (expected UUID)")
    return bid


_lock = threading.Lock()
_store: dict[str, dict[str, Any]] = {}


def storage_mode() -> str:
    return "disk+memory" if (os.environ.get("BATCH_STORE_PATH", "").strip()) else "memory"


def _disk_store_dir() -> Path | None:
    raw = os.environ.get("BATCH_STORE_PATH", "").strip()
    if not raw:
        return None
    p = Path(raw).expanduser()
    try:
        p.mkdir(parents=True, exist_ok=True)
    except Exception:
        return None
    return p


def _disk_record_path(batch_id: str) -> Path | None:
    root = _disk_store_dir()
    if root is None:
        return None
    return root / f"{batch_id}.json"


def _write_disk_record(rec: dict[str, Any]) -> None:
    p = _disk_record_path(str(rec.get("batch_id", "")))
    if p is None:
        return
    payload = {
        "batch_id": rec.get("batch_id"),
        "created_at": float(rec.get("created_at", time.time())),
        "tenant_id": rec.get("tenant_id"),
        "analyst_id": rec.get("analyst_id"),
        "filename": rec.get("filename"),
        "format": rec.get("format"),
        "columns": rec.get("columns") or [],
        "rows": rec.get("rows") or [],
        "row_count": int(rec.get("row_count") or 0),
    }
    p.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")


def _read_disk_record(batch_id: str) -> dict[str, Any] | None:
    p = _disk_record_path(batch_id)
    if p is None or not p.exists():
        return None
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(raw, dict):
        return None
    rows = raw.get("rows")
    cols = raw.get("columns")
    if not isinstance(rows, list) or not isinstance(cols, list):
        return None
    return {
        "batch_id": str(raw.get("batch_id") or batch_id),
        "created_at": float(raw.get("created_at") or 0.0),
        "tenant_id": str(raw.get("tenant_id") or ""),
        "analyst_id": str(raw.get("analyst_id") or ""),
        "filename": str(raw.get("filename") or "upload"),
        "format": str(raw.get("format") or "json"),
        "columns": [str(c)[:256] for c in cols[:_MAX_COLS]],
        "rows": rows[:_MAX_ROWS],
        "row_count": int(raw.get("row_count") or len(rows)),
    }


def _cleanup_disk(now: float) -> None:
    root = _disk_store_dir()
    if root is None:
        return
    files = sorted(root.glob("*.json"), key=lambda p: p.stat().st_mtime if p.exists() else 0.0)
    ttl = ttl_seconds()
    alive: list[Path] = []
    for p in files:
        try:
            rec = json.loads(p.read_text(encoding="utf-8"))
            created_at = float(rec.get("created_at") or p.stat().st_mtime)
        except Exception:
            try:
                p.unlink(missing_ok=True)
            except Exception:
                pass
            continue
        if now - created_at > ttl:
            try:
                p.unlink(missing_ok=True)
            except Exception:
                pass
            continue
        alive.append(p)
    if len(alive) <= _MAX_STORE_BATCHES:
        return
    for p in alive[: max(0, len(alive) - _MAX_STORE_BATCHES)]:
        try:
            p.unlink(missing_ok=True)
        except Exception:
            pass


def _cleanup_unlocked(now: float) -> None:
    dead = [bid for bid, rec in _store.items() if now - float(rec.get("created_at", 0)) > ttl_seconds()]
    for bid in dead:
        del _store[bid]
    _cleanup_disk(now)
    if len(_store) <= _MAX_STORE_BATCHES:
        return
    # Drop oldest beyond cap
    ordered = sorted(_store.items(), key=lambda x: float(x[1].get("created_at", 0)))
    for bid, _ in ordered[: max(0, len(_store) - _MAX_STORE_BATCHES)]:
        del _store[bid]


def _normalize_key(k: str) -> str:
    s = str(k).strip()
    return s[:256] if s else "column"


def _parse_csv(raw: bytes) -> tuple[list[str], list[dict[str, Any]], str]:
    text = raw.decode("utf-8-sig", errors="replace")
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    if not reader.fieldnames:
        raise ValueError("CSV has no header row")
    columns = [_normalize_key(c) for c in reader.fieldnames[:_MAX_COLS]]
    rows: list[dict[str, Any]] = []
    for i, row in enumerate(reader):
        if i >= _MAX_ROWS:
            break
        out: dict[str, Any] = {}
        for c in columns:
            v = row.get(c)
            out[c] = "" if v is None else str(v)[:4096]
        rows.append(out)
    return columns, rows, "csv"


def _parse_json(raw: bytes) -> tuple[list[str], list[dict[str, Any]], str]:
    text = raw.decode("utf-8-sig", errors="replace")
    data = json.loads(text)
    rows_in: list[dict[str, Any]] = []

    if isinstance(data, list):
        for x in data:
            if isinstance(x, dict):
                rows_in.append(x)
    elif isinstance(data, dict):
        for key in ("items", "data", "rows", "records"):
            v = data.get(key)
            if isinstance(v, list) and v and isinstance(v[0], dict):
                rows_in = [r for r in v if isinstance(r, dict)]
                break
        if not rows_in:
            # column-oriented: { "a": [1,2], "b": [3,4] }
            lists = [(k, v) for k, v in data.items() if isinstance(v, list)]
            if lists:
                n = min(len(v) for _, v in lists)
                n = min(n, _MAX_ROWS)
                keys = [_normalize_key(k) for k, _ in lists[:_MAX_COLS]]
                for i in range(n):
                    row = {}
                    for (k, v), kn in zip(lists[:_MAX_COLS], keys):
                        row[kn] = v[i]
                    rows_in.append(row)

    if not rows_in:
        raise ValueError("JSON must be an array of objects, an object with items/data/rows, or column-oriented arrays")

    rows_in = rows_in[:_MAX_ROWS]
    colset: list[str] = []
    seen: set[str] = set()
    for r in rows_in:
        for k in r.keys():
            nk = _normalize_key(str(k))
            if nk not in seen:
                seen.add(nk)
                colset.append(nk)
                if len(colset) >= _MAX_COLS:
                    break
        if len(colset) >= _MAX_COLS:
            break

    rows: list[dict[str, Any]] = []
    for r in rows_in:
        out: dict[str, Any] = {}
        for c in colset:
            v = r.get(c)
            if v is None and c not in r:
                # original key might differ only by case — try first match
                for rk in r:
                    if _normalize_key(str(rk)) == c:
                        v = r[rk]
                        break
            if isinstance(v, (dict, list)):
                out[c] = json.dumps(v, default=str)[:4096]
            elif v is None:
                out[c] = ""
            else:
                out[c] = v
        rows.append(out)

    return colset, rows, "json"


def _parse_xlsx(raw: bytes) -> tuple[list[str], list[dict[str, Any]], str]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as e:
        raise ValueError("Excel support requires openpyxl (install investigation-agent with excel extra)") from e

    bio = io.BytesIO(raw)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("Excel file is empty")
        columns = []
        for c in header_row[:_MAX_COLS]:
            if c is None:
                columns.append(f"column_{len(columns)}")
            else:
                columns.append(_normalize_key(str(c)))
        if not any(columns):
            raise ValueError("Excel has no header row")
        rows: list[dict[str, Any]] = []
        for i, row in enumerate(rows_iter):
            if i >= _MAX_ROWS:
                break
            d: dict[str, Any] = {}
            for j, col in enumerate(columns):
                v = row[j] if j < len(row) else None
                if v is None:
                    d[col] = ""
                elif isinstance(v, float) and v == int(v):
                    d[col] = int(v)
                else:
                    d[col] = v
            rows.append(d)
        return columns, rows, "xlsx"
    finally:
        wb.close()


def parse_tabular_file(filename: str, raw: bytes) -> tuple[list[str], list[dict[str, Any]], str]:
    if len(raw) > _MAX_FILE_BYTES:
        raise ValueError(f"File exceeds max size ({_MAX_FILE_BYTES // (1024 * 1024)} MiB)")
    name = (filename or "upload").lower()
    if name.endswith(".csv"):
        return _parse_csv(raw)
    if name.endswith(".ndjson"):
        lines = raw.decode("utf-8-sig", errors="replace").splitlines()
        objs: list[dict[str, Any]] = []
        for line in lines[:_MAX_ROWS]:
            line = line.strip()
            if not line:
                continue
            o = json.loads(line)
            if isinstance(o, dict):
                objs.append(o)
        if not objs:
            raise ValueError("NDJSON contained no JSON objects")
        return _parse_json(json.dumps(objs).encode())
    if name.endswith(".json"):
        return _parse_json(raw)
    if name.endswith(".xlsx"):
        return _parse_xlsx(raw)
    if name.endswith(".xls"):
        raise ValueError("Legacy .xls is not supported; save as .xlsx or use CSV")
    raise ValueError("Unsupported format; use .csv, .json, .ndjson, or .xlsx")


def store_batch(
    tenant_id: str,
    analyst_id: str,
    filename: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    format_name: str,
) -> str:
    batch_id = str(uuid.uuid4())
    now = time.time()
    with _lock:
        _cleanup_unlocked(now)
        rec = {
            "batch_id": batch_id,
            "created_at": now,
            "tenant_id": tenant_id,
            "analyst_id": analyst_id,
            "filename": filename[:512],
            "format": format_name,
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
        }
        _store[batch_id] = rec
        _write_disk_record(rec)
    return batch_id


def get_batch(batch_id: str, tenant_id: str, analyst_id: str) -> dict[str, Any] | None:
    try:
        validate_batch_id(batch_id)
    except ValueError:
        return None
    now = time.time()
    with _lock:
        _cleanup_unlocked(now)
        rec = _store.get(batch_id)
        if not rec:
            rec = _read_disk_record(batch_id)
            if rec:
                _store[batch_id] = rec
        if not rec:
            return None
        if rec.get("tenant_id") != tenant_id or rec.get("analyst_id") != analyst_id:
            return None
        return rec


def batch_profile(rec: dict[str, Any]) -> dict[str, Any]:
    columns = rec.get("columns") or []
    rows: list[dict[str, Any]] = rec.get("rows") or []
    col_info: list[dict[str, Any]] = []
    for c in columns:
        non_null = 0
        numeric_try = 0
        numeric_ok = 0
        for r in rows[:500]:
            v = r.get(c)
            if v is not None and v != "":
                non_null += 1
                if isinstance(v, (int, float)):
                    numeric_ok += 1
                    numeric_try += 1
                else:
                    numeric_try += 1
                    try:
                        float(str(v).replace(",", ""))
                        numeric_ok += 1
                    except ValueError:
                        pass
        ratio = (numeric_ok / numeric_try) if numeric_try else 0.0
        inferred = "numeric" if ratio > 0.85 else "string"
        col_info.append(
            {
                "column": c,
                "non_null_sample_of_500": non_null,
                "inferred_type": inferred,
            }
        )
    sample = rows[:5]
    return {
        "batch_id": rec.get("batch_id", ""),
        "filename": rec.get("filename"),
        "format": rec.get("format"),
        "row_count": rec.get("row_count"),
        "columns": columns,
        "column_profiles": col_info,
        "sample_rows": sample,
    }


def batch_query_rows(rec: dict[str, Any], offset: int, limit: int, columns: list[str] | None) -> dict[str, Any]:
    rows: list[dict[str, Any]] = rec.get("rows") or []
    all_cols: list[str] = rec.get("columns") or []
    off = max(0, offset)
    lim = max(1, min(limit, 100))
    slice_rows = rows[off : off + lim]
    use_cols = columns if columns else all_cols
    safe_cols = [c for c in use_cols if c in all_cols][:_MAX_COLS]
    if not safe_cols:
        safe_cols = all_cols
    trimmed = [{k: r.get(k) for k in safe_cols} for r in slice_rows]
    return {
        "offset": off,
        "limit": lim,
        "returned": len(trimmed),
        "total_rows": len(rows),
        "columns": safe_cols,
        "rows": trimmed,
    }


def batch_aggregate_column(rec: dict[str, Any], column: str, mode: str) -> dict[str, Any]:
    all_cols: list[str] = rec.get("columns") or []
    if column not in all_cols:
        return {"error": "unknown_column", "column": column, "known": all_cols}
    rows: list[dict[str, Any]] = rec.get("rows") or []
    mode = (mode or "value_counts").lower()
    vals = [r.get(column) for r in rows]

    if mode == "numeric_summary":
        nums: list[float] = []
        for v in vals:
            if v is None or v == "":
                continue
            if isinstance(v, (int, float)):
                nums.append(float(v))
            else:
                try:
                    nums.append(float(str(v).replace(",", "")))
                except ValueError:
                    continue
        if not nums:
            return {"column": column, "mode": mode, "error": "no_numeric_values"}
        nums.sort()
        s = sum(nums)
        return {
            "column": column,
            "mode": mode,
            "count": len(nums),
            "min": nums[0],
            "max": nums[-1],
            "mean": round(s / len(nums), 6),
        }

    # value_counts (default)
    strs = ["" if v is None else str(v) for v in vals]
    ctr = Counter(strs)
    top = ctr.most_common(25)
    return {
        "column": column,
        "mode": "value_counts",
        "distinct": len(ctr),
        "top_values": [{"value": a[:500], "count": b} for a, b in top],
    }
