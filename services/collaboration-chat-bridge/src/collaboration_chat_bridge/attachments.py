from __future__ import annotations

"""Download Slack files and extract text for copilot context (PDF, plain text, CSV, xlsx)."""


import logging
from io import BytesIO
from typing import Any

import httpx

logger = logging.getLogger(__name__)

_TEXT_PREFIXES = ("text/", "application/json", "application/xml")


def extract_text_from_bytes(data: bytes, *, filename: str, mimetype: str) -> str:
    """Best-effort text extraction; caps applied by caller."""
    mt = (mimetype or "").split(";")[0].strip().lower()
    name = (filename or "file").lower()

    if mt.startswith("text/") or mt in ("application/json", "application/xml", "application/csv"):
        return data.decode("utf-8", errors="replace")

    if "pdf" in mt or name.endswith(".pdf"):
        try:
            from pypdf import PdfReader
        except ImportError as e:  # pragma: no cover
            raise RuntimeError("pypdf is required for PDF attachments") from e
        reader = PdfReader(BytesIO(data))
        parts: list[str] = []
        for page in reader.pages:
            try:
                parts.append(page.extract_text() or "")
            except Exception as ex:  # noqa: BLE001
                logger.warning("pypdf page extract failed: %s", ex)
                parts.append("")
        return "\n".join(parts).strip()

    if "csv" in mt or name.endswith(".csv"):
        return data.decode("utf-8", errors="replace")

    if "spreadsheetml" in mt or name.endswith(".xlsx"):
        return _extract_xlsx_text(data)

    return f"(unsupported attachment type: {mimetype or 'unknown'}; {len(data)} bytes)"


def _extract_xlsx_text(data: bytes) -> str:
    """Read first sheets/rows as tab-separated text (bounded for memory)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel attachments") from e

    max_sheets, max_rows, max_cols = 5, 500, 50
    buf = BytesIO(data)
    wb = load_workbook(buf, read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for idx, sheet in enumerate(wb.worksheets):
            if idx >= max_sheets:
                parts.append("\n…(further sheets omitted)\n")
                break
            lines: list[str] = [f"\n--- Sheet: {sheet.title} ---\n"]
            row_count = 0
            for row in sheet.iter_rows(
                min_row=1,
                max_row=max_rows,
                max_col=max_cols,
                values_only=True,
            ):
                row_count += 1
                cells = []
                for val in row:
                    if val is None:
                        cells.append("")
                    else:
                        s = str(val).replace("\t", " ").replace("\n", " ").strip()
                        cells.append(s)
                lines.append("\t".join(cells))
            if row_count >= max_rows:
                lines.append(f"…(truncated after {max_rows} rows)\n")
            parts.append("\n".join(lines))
        return "\n".join(parts).strip()
    finally:
        wb.close()


async def download_slack_file_private(url: str, bot_token: str, *, max_bytes: int) -> bytes:
    headers = {"Authorization": f"Bearer {bot_token}"}
    async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
        r = await client.get(url, headers=headers)
        r.raise_for_status()
        data = r.content
    if len(data) > max_bytes:
        return data[:max_bytes]
    return data


async def slack_files_to_append_text(
    files: list[dict[str, Any]],
    bot_token: str,
    *,
    max_bytes_per_file: int,
    max_total_chars: int,
) -> str:
    """Download Slack `files` entries and return a single appendix block."""
    if not files or not bot_token.strip():
        return ""
    chunks: list[str] = []
    total = 0
    for f in files[:5]:
        if not isinstance(f, dict):
            continue
        url = (f.get("url_private_download") or f.get("url_private") or "").strip()
        if not url:
            continue
        name = str(f.get("name") or "attachment")
        mimetype = str(f.get("mimetype") or "application/octet-stream")
        try:
            raw = await download_slack_file_private(url, bot_token, max_bytes=max_bytes_per_file)
            text = extract_text_from_bytes(raw, filename=name, mimetype=mimetype)
        except Exception as e:  # noqa: BLE001
            logger.warning("slack file download/extract failed: %s", e)
            chunks.append(f"\n[Attachment {name}: could not read ({e})]\n")
            continue
        if len(text) > 24_000:
            text = text[:24_000] + "\n…(truncated)"
        block = f"\n\n--- Attachment: {name} ({mimetype}) ---\n{text}\n"
        if total + len(block) > max_total_chars:
            chunks.append("\n…(further attachments omitted: size cap)\n")
            break
        chunks.append(block)
        total += len(block)
    return "".join(chunks)
